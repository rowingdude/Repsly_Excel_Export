#
#        Repsly2Excel
#
#        Author: Benjamin Cance
#
#        Email: bjc@tdx.li
#
#        Website: https://tdx.li
#
#        License: MIT License
#
#        Copyright © 2024 - Benjamin Cance
#        
#        Permission is hereby granted, free of charge, to any person obtaining a copy
#        of this software and associated documentation files (the "Software"), to deal
#        in the Software without restriction, including without limitation the rights
#        to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
#        copies of the Software, and to permit persons to whom the Software is
#        furnished to do so, subject to the following conditions:
#
#        The above copyright notice and this permission notice shall be included in all
#        copies or substantial portions of the Software.
#
#        THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
#        IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
#        FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
#        AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
#        LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
#        OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
#        SOFTWARE.
#

import asyncio
import aiohttp
import aiostream     # Future thoughts
import argparse
import base64
import functools
import itertools     # Not used but here because of future ideas
import json
import logging
import os
import requests      # Not used but here because of future ideas
import sys
import time

from datetime import datetime, timedelta
from logging.handlers import RotatingFileHandler
from openpyxl import Workbook, load_workbook

BASE_URL = "https://api.repsly.com/v3/export"
API_USERNAME = os.environ.get('REPSLY_API_USERNAME')
API_PASSWORD = os.environ.get('REPSLY_API_PASSWORD')

if not API_USERNAME or not API_PASSWORD:
    raise ValueError("API credentials not set. Please set REPSLY_API_USERNAME and REPSLY_API_PASSWORD environment variables.")

auth_header = base64.b64encode(f"{API_USERNAME}:{API_PASSWORD}".encode()).decode()
headers = {
    "Authorization": f"Basic {auth_header}",
    "Content-Type": "application/json"
}

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

def setup_logging():
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)

    file_handler = RotatingFileHandler('repsly_export.log', maxBytes=10*1024*1024, backupCount=5)
    file_handler.setLevel(logging.DEBUG)
    file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(file_formatter)

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_formatter = logging.Formatter('%(message)s')
    console_handler.setFormatter(console_formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger

logger = setup_logging()

def log_function_call(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        
        if func.__name__ == 'fetch_data':
            url = args[0] if args else kwargs.get('url')
            if not result:
                logger.error(f"Error fetching data from {url}")
                response = kwargs.get('response')
                if response:
                    logger.error(f"Status code: {response.status_code}")
                    logger.error(f"Headers: {response.headers}")
                    logger.error(f"Content: {response.text}")
            else:
                logger.debug(f"Successfully fetched data from {url}")
        else:
            logger.debug(f"Finished {func.__name__}. Execution time: {end_time - start_time:.2f} seconds")

        if isinstance(result, str) and result.endswith('.xlsx'):
            try:
                wb = load_workbook(result)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    logger.debug(f"Sheet '{sheet_name}' in {result} has {sheet.max_row} rows and {sheet.max_column} columns")
                    if sheet.max_row > 1:
                        first_data_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
                        logger.debug(f"First data row in '{sheet_name}': {first_data_row}")
                    else:
                        logger.warning(f"Sheet '{sheet_name}' in {result} has no data rows")
            except Exception as e:
                logger.error(f"Error analyzing {result}: {str(e)}")
        return result
    return wrapper

def save_last_ids(last_ids, filename='last_ids.json'):
    with open(filename, 'w') as f:
        json.dump(last_ids, f, indent=2)

def load_last_ids(filename='last_ids.json'):
    if os.path.exists(filename):
        with open(filename, 'r') as f:
            return json.load(f)
    return {}


async def fetch_data(session, url, params=None):
    try:
        async with session.get(url, headers=headers, params=params) as response:
            if response.status == 200:
                return await response.json(), response
            else:
                logger.error(f"Error fetching data from {url}: {response.status}")
                return None, response
    except Exception as e:
        logger.error(f"Exception occurred while fetching data from {url}: {str(e)}")
        return None, None

def process_field(value):
    if isinstance(value, list):
        return ', '.join(str(v) for v in value)
    elif isinstance(value, dict):
        return ', '.join(f"{k}:{v}" for k, v in value.items())
    return value

async def process_data_async(session, endpoint, key_name, headers, last_value=0, use_timestamp=False):
    filename = f"Repsly_{key_name}_Export.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = key_name
    ws.append(headers)
    row_count = 0
    save_interval = 50  # Save every 1000 rows

    while True:
        url = f"{BASE_URL}/{endpoint}/{last_value}"
        data, _ = await fetch_data(session, url)
        
        if data and key_name in data:
            items = data[key_name]
            for item in items:
                row = [process_field(item.get(header)) for header in headers]
                ws.append(row)
                row_count += 1

            if row_count % save_interval == 0:
                wb.save(filename)
                logger.debug(f"Saved progress for {key_name}. Current row count: {row_count}")

            meta = data.get('MetaCollectionResult', {})
            if use_timestamp:
                new_last_value = meta.get('LastTimeStamp')
            else:
                new_last_value = meta.get('LastID')
            
            if new_last_value is not None and new_last_value != last_value:
                last_value = new_last_value
            else:
                break  # No more data to fetch

            if len(items) < 50:
                break
        else:
            logger.warning(f"No '{key_name}' found in data from {url}")
            break

    wb.save(filename)
    logger.info(f"{key_name} data saved to {filename}. Total rows: {row_count}")
    return filename, last_value


async def process_clients(session, last_id=0):
    headers = [
        "ClientID", "TimeStamp", "Code", "Name", "Active", "Tag", "Territory",
        "RepresentativeCode", "RepresentativeName", "StreetAddress", "ZIP", "City",
        "State", "Country", "Email", "Phone", "Mobile", "Website", "ContactName",
        "ContactTitle", "Note", "Status", "CustomFields", "PriceLists", "AccountCode"
    ]
    return await process_data_async(session, "clients", "Clients", headers, last_id, use_timestamp=False)


async def process_client_notes(session, last_id=0):
    headers = [
        "ClientNoteID", "TimeStamp", "DateAndTime", "RepresentativeCode",
        "RepresentativeName", "ClientCode", "ClientName", "StreetAddress",
        "ZIP", "ZIPExt", "City", "State", "Country", "Email", "Phone",
        "Mobile", "Territory", "Longitude", "Latitude", "Note", "VisitID"
    ]
    return await process_data_async(session, "clientnotes", "ClientNotes", headers, last_id, use_timestamp=False)


async def process_visits(session, last_timestamp=0):
    headers = [
        "VisitID", "TimeStamp", "Date", "RepresentativeCode", "RepresentativeName",
        "ExplicitCheckIn", "DateAndTimeStart", "DateAndTimeEnd", "ClientCode",
        "ClientName", "StreetAddress", "ZIP", "ZIPExt", "City", "State", "Country",
        "Territory", "LatitudeStart", "LongitudeStart", "LatitudeEnd", "LongitudeEnd",
        "PrecisionStart", "PrecisionEnd", "VisitStatusBySchedule", "VisitEnded"
    ]
    return await process_data_async(session, "visits", "Visits", headers, last_timestamp, use_timestamp=True)


async def process_retail_audits(session, last_id=0):
    headers = [
        "RetailAuditID", "RetailAuditName", "Cancelled", "ClientCode", "ClientName",
        "DateAndTime", "RepresentativeCode", "RepresentativeName", "ProductGroupCode",
        "ProductGroupName", "ProductCode", "ProductName", "Present", "Price",
        "Promotion", "ShelfShare", "ShelfSharePercent", "SoldOut", "Stock",
        "CustomFields", "Note", "VisitID"
    ]
    return await process_data_async(session, "retailaudits", "RetailAudits", headers, last_id, use_timestamp=False)


async def process_purchase_orders(session, last_id=0):
    headers = [
        "PurchaseOrderID", "TransactionType", "DocumentTypeID", "DocumentTypeName",
        "DocumentStatus", "DocumentStatusID", "DocumentItemAttributeCaption",
        "DateAndTime", "DocumentNo", "ClientCode", "ClientName", "DocumentDate",
        "DueDate", "RepresentativeCode", "RepresentativeName", "LineNo",
        "ProductCode", "ProductName", "UnitAmount", "UnitPrice", "PackageTypeCode",
        "PackageTypeName", "PackageTypeConversion", "Quantity", "Amount",
        "DiscountAmount", "DiscountPercent", "TaxAmount", "TaxPercent", "TotalAmount",
        "ItemNote", "DocumentItemAttributeName", "DocumentItemAttributeID",
        "SignatureURL", "Note", "Taxable", "VisitID", "StreetAddress", "ZIP",
        "ZIPExt", "City", "State", "Country", "CountryCode", "CustomAttributes",
        "OriginalDocumentNumber"
    ]
    return await process_data_async(session, "purchaseorders", "PurchaseOrders", headers, last_id, use_timestamp=False)


async def process_products(session, last_id=0):
    headers = [
        "Code", "Name", "ProductGroupCode", "ProductGroupName", "Active", "Tag",
        "UnitPrice", "EAN", "Note", "ImageUrl", "MasterProduct", "PackagingCodes"
    ]
    return await process_data_async(session, "products", "Products", headers, last_id, use_timestamp=False)


async def process_forms(session, last_id=0):
    headers = [
        "FormID", "FormName", "ClientCode", "ClientName", "DateAndTime",
        "RepresentativeCode", "RepresentativeName", "StreetAddress", "ZIP",
        "ZIPExt", "City", "State", "Country", "Email", "Phone", "Mobile",
        "Territory", "Longitude", "Latitude", "SignatureURL", "VisitStart",
        "VisitEnd", "VisitID", "FormItems"
    ]
    return await process_data_async(session, "forms", "Forms", headers, last_id, use_timestamp=False)


async def process_photos(session, last_id=0):
    headers = [
        "PhotoID", "ClientCode", "ClientName", "Note", "DateAndTime", "PhotoURL",
        "RepresentativeCode", "RepresentativeName", "VisitID", "Tag"
    ]
    return await process_data_async(session, "photos", "Photos", headers, last_id, use_timestamp=False)


async def process_daily_working_time(session, last_id=0):
    headers = [
        "DailyWorkingTimeID", "Date", "DateAndTimeStart", "DateAndTimeEnd",
        "Length", "MileageStart", "MileageEnd", "MileageTotal", "LatitudeStart",
        "LongitudeStart", "LatitudeEnd", "LongitudeEnd", "RepresentativeCode",
        "RepresentativeName", "Note", "Tag", "NoOfVisits", "MinOfVisits",
        "MaxOfVisits", "MinMaxVisitsTime", "TimeAtClient", "TimeAtTravel"
    ]
    return await process_data_async(session, "dailyworkingtime", "DailyWorkingTime", headers, last_id, use_timestamp=False)


async def process_visit_schedules(session, last_id=None):
    headers = [
        "ScheduleDateAndTime", "RepresentativeCode", "RepresentativeName",
        "ClientCode", "ClientName", "StreetAddress", "ZIP", "ZIPExt", "City",
        "State", "Country", "Territory", "VisitNote", "DueDate"
    ]
    
    filename = "Repsly_VisitSchedules_Export.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "VisitSchedules"
    ws.append(headers)

    end_date = datetime.now().strftime("%Y-%m-%d")
    begin_date = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    url = f"{BASE_URL}/visitschedules/{begin_date}/{end_date}"
    
    row_count = 0
    save_interval = 1000

    while True:
        data, _ = await fetch_data(session, url)
        
        if data and 'VisitSchedules' in data:
            for schedule in data['VisitSchedules']:
                ws.append([process_field(schedule.get(header)) for header in headers])
                row_count += 1

            if row_count % save_interval == 0:
                wb.save(filename)
                logger.debug(f"Saved progress for Visit Schedules. Current row count: {row_count}")

            if len(data['VisitSchedules']) < 50:
                break
            
            last_schedule = data['VisitSchedules'][-1]
            begin_date = last_schedule['ScheduleDateAndTime'].split('T')[0]
            url = f"{BASE_URL}/visitschedules/{begin_date}/{end_date}"
        else:
            break

    wb.save(filename)
    logging.info(f"Visit Schedules data saved to {filename}. Total rows: {row_count}")
    return filename, None

async def process_visit_realizations(session, last_id=None):
    headers = [
        "ScheduleId", "ProjectId", "EmployeeId", "EmployeeCode", "PlaceId",
        "PlaceCode", "ModifiedUTC", "TimeZone", "ScheduleNote", "Status",
        "DateTimeStart", "DateTimeStartUTC", "DateTimeEnd", "DateTimeEndUTC",
        "PlanDateTimeStart", "PlanDateTimeStartUTC", "PlanDateTimeEnd",
        "PlanDateTimeEndUTC", "Tasks"
    ]
    
    filename = "Repsly_VisitRealizations_Export.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "VisitRealizations"
    ws.append(headers)

    modified_date = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%dT%H:%M:%S.000Z")
    skip = 0
    row_count = 0
    save_interval = 1000

    while True:
        url = f"{BASE_URL}/visitrealizations"
        params = {'modified': modified_date, 'skip': skip}
        data, _ = await fetch_data(session, url, params)

        if data and 'VisitRealizations' in data:
            for visit in data['VisitRealizations']:
                ws.append([process_field(visit.get(header)) for header in headers])
                row_count += 1
            
            if row_count % save_interval == 0:
                wb.save(filename)
                logger.debug(f"Saved progress for Visit Realizations. Current row count: {row_count}")
            
            if len(data['VisitRealizations']) < 50:
                break
            skip += 50
        else:
            break

    wb.save(filename)
    logging.info(f"Visit Realizations data saved to {filename}. Total rows: {row_count}")
    return filename, None

async def process_representatives(session, last_id=None):
    headers = [
        "Code", "Name", "Note", "Email", "Phone", "Territories", "Active",
        "Address1", "Address2", "City", "State", "ZipCode", "ZipCodeExt",
        "Country", "CountryCode", "Attributes"
    ]
    filename = "Repsly_Representatives_Export.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Representatives"
    ws.append(headers)

    url = f"{BASE_URL}/representatives"
    row_count = 0
    save_interval = 1000

    while True:
        data, _ = await fetch_data(session, url)
        
        if data and 'Representatives' in data:
            for rep in data['Representatives']:
                try:
                    attributes = rep.get('Attributes', [])
                    if attributes is None:
                        attributes = []
                    attributes_str = ', '.join([f"{attr.get('Title', '')}:{attr.get('Type', '')}:{attr.get('Value', '')}" for attr in attributes])
                    
                    ws.append([
                        rep.get('Code'),
                        rep.get('Name'),
                        rep.get('Note'),
                        rep.get('Email'),
                        rep.get('Phone'),
                        ', '.join(rep.get('Territories', [])),
                        rep.get('Active'),
                        rep.get('Address1'),
                        rep.get('Address2'),
                        rep.get('City'),
                        rep.get('State'),
                        rep.get('ZipCode'),
                        rep.get('ZipCodeExt'),
                        rep.get('Country'),
                        rep.get('CountryCode'),
                        attributes_str
                    ])
                    row_count += 1

                    if row_count % save_interval == 0:
                        wb.save(filename)
                        logger.debug(f"Saved progress for Representatives. Current row count: {row_count}")

                except Exception as e:
                    logging.error(f"Error processing representative: {rep.get('Code', 'Unknown')}. Error: {str(e)}")
            
            if len(data['Representatives']) < 50:
                break
            
        else:
            logging.warning("No 'Representatives' data found in the API response")
            break

    wb.save(filename)
    logging.info(f"Representatives data saved to {filename}. Total rows: {row_count}")
    return filename, None

async def process_users(session, last_timestamp=0):
    headers = [
        "ID", "Code", "Name", "Email", "Active", "Role", "Note", "Phone",
        "Territories", "SendEmailEnabled", "Address1", "Address2", "City",
        "State", "ZipCode", "ZipCodeExt", "Country", "CountryCode",
        "Attributes", "Permissions"
    ]
    return await process_data_async(session, "users", "Users", headers, last_timestamp, use_timestamp=True)


async def process_document_types(session, last_id=None):
    headers = ["DocumentTypeID", "DocumentTypeName", "Statuses", "Pricelists"]
    filename = "Repsly_DocumentTypes_Export.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "DocumentTypes"
    ws.append(headers)

    url = f"{BASE_URL}/documentTypes"
    data, _ = await fetch_data(session, url)
    row_count = 0
    
    if data and 'DocumentTypes' in data:
        for doc_type in data['DocumentTypes']:
            ws.append([
                doc_type.get('DocumentTypeID'),
                doc_type.get('DocumentTypeName'),
                ', '.join([status.get('DocumentStatusName', '') for status in doc_type.get('Statuses', [])]),
                ', '.join([pricelist.get('PricelistName', '') for pricelist in doc_type.get('Pricelists', [])])
            ])
            row_count += 1

    wb.save(filename)
    logging.info(f"Document Types data saved to {filename}. Total rows: {row_count}")
    return filename, None


async def process_pricelists(session, last_id=None):
    headers = ["ID", "Name", "IsDefault", "Active", "UsePrices"]
    filename = "Repsly_Pricelists_Export.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Pricelists"
    ws.append(headers)

    url = f"{BASE_URL}/pricelists"
    row_count = 0
    save_interval = 1000
    
    while True:
        data, _ = await fetch_data(session, url)
        
        if data and 'Pricelists' in data:
            for pricelist in data['Pricelists']:
                ws.append([
                    pricelist.get('ID'),
                    pricelist.get('Name'),
                    pricelist.get('IsDefault'),
                    pricelist.get('Active'),
                    pricelist.get('UsePrices')
                ])
                row_count += 1

                if row_count % save_interval == 0:
                    wb.save(filename)
                    logger.debug(f"Saved progress for Pricelists. Current row count: {row_count}")

            if len(data['Pricelists']) < 50:
                break
            
        else:
            break

    wb.save(filename)
    logging.info(f"Pricelists data saved to {filename}. Total rows: {row_count}")
    return filename, None

async def process_pricelist_items(session,last_id=None):
    headers = [
        "PricelistID", "ID", "ProductID", "ProductCode", "Price", "Active",
        "ClientID", "ManufactureID", "DateAvailableFrom", "DateAvailableTo",
        "MinQuantity", "MaxQuantity"
    ]
    
    filename = "Repsly_PricelistItems_Export.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "PricelistItems"
    ws.append(headers)

    row_count = 0
    pricelists_url = f"{BASE_URL}/pricelists"
    pricelists_data, _ = await fetch_data(session, pricelists_url)
    
    if pricelists_data and 'Pricelists' in pricelists_data:
        for pricelist in pricelists_data['Pricelists']:
            pricelist_id = pricelist.get('ID')
            if pricelist_id:
                url = f"{BASE_URL}/pricelistsItems/{pricelist_id}"
                data, _ = await fetch_data(session, url)
                
                if data and isinstance(data, list):
                    for item in data:
                        ws.append([
                            pricelist_id,
                            item.get('ID'),
                            item.get('ProductID'),
                            item.get('ProductCode'),
                            item.get('Price'),
                            item.get('Active'),
                            item.get('ClientID'),
                            item.get('ManufactureID'),
                            item.get('DateAvailableFrom'),
                            item.get('DateAvailableTo'),
                            item.get('MinQuantity'),
                            item.get('MaxQuantity')
                        ])
                        row_count += 1

    wb.save(filename)
    logging.info(f"Pricelist Items data saved to {filename}. Total rows: {row_count}")
    return filename, None

async def create_combined_workbook(filenames):
    combined_wb = Workbook()
    combined_wb.remove(combined_wb.active)  

    for filename in filenames:
        if os.path.exists(filename):
            try:
                wb = load_workbook(filename)
                for sheet_name in wb.sheetnames:
                    source_sheet = wb[sheet_name]
                    new_sheet = combined_wb.create_sheet(sheet_name)
                    
                    for row in source_sheet.iter_rows(values_only=True):
                        new_sheet.append(row)
                
                os.remove(filename)
            except Exception as e:
                logging.error(f"Error processing file {filename}: {str(e)}")
        else:
            logging.warning(f"File not found: {filename}")

    if not combined_wb.sheetnames:
        logging.warning("No data was combined. Creating a default sheet.")
        combined_wb.create_sheet("Empty")

    return combined_wb

async def process_import_status(session, import_job_id):
    headers = [
        "ImportStatus", "RowsInserted", "RowsUpdated", "RowsInvalid", "RowsTotal",
        "Warnings", "Errors"
    ]
    
    filename = "Repsly_ImportStatus_Export.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "ImportStatus"
    ws.append(headers)

    url = f"{BASE_URL}/importStatus/{import_job_id}"
    data, _ = await fetch_data(session, url)
    
    if data:
        warnings = '; '.join([f"{w['ItemID']}:{w['ItemName']}:{w['ItemStatus']}" for w in data.get('Warnings', [])])
        errors = '; '.join([f"{e['ItemID']}:{e['ItemName']}:{e['ItemStatus']}" for e in data.get('Errors', [])])
        
        ws.append([
            data.get('ImportStatus'),
            data.get('RowsInserted'),
            data.get('RowsUpdated'),
            data.get('RowsInvalid'),
            data.get('RowsTotal'),
            warnings,
            errors
        ])

    wb.save(filename)
    logging.info(f"Import Status data saved to {filename}")
    return filename

async def main(modules=None):
    last_ids = load_last_ids()
    filenames = []

    all_endpoints = {
        'representatives': process_representatives,
        'visitschedules': process_visit_schedules,
        'pricelistitems': process_pricelist_items,
        'pricelists': process_pricelists,
        'documenttypes': process_document_types,
        'purchaseorders': process_purchase_orders,
        'clients': process_clients,
        'clientnotes': process_client_notes,
        'visits': process_visits,
        'retailaudits': process_retail_audits,
        'products': process_products,
        'forms': process_forms,
        'photos': process_photos,
        'dailyworkingtime': process_daily_working_time,
        'visitrealizations': process_visit_realizations,
        'users': process_users,
    }

    if not modules:
        modules = all_endpoints.keys()

    logger.info("Starting Repsly data export...")

    async def process_module(session, module):
        if module in all_endpoints:
            logger.info(f"Processing {module}...")
            process_func = all_endpoints[module]
            try:
                last_id = last_ids.get(module, 0)
                filename, new_last_id = await process_func(session, last_id)
                if filename:
                    logger.info(f"{module.capitalize()} data exported successfully.")
                return module, filename, new_last_id
            except Exception as e:
                logger.error(f"Error processing {module}: {str(e)}", exc_info=True)
                return module, None, None
        else:
            logger.warning(f"Unknown module: {module}")
            return module, None, None
    

    async with aiohttp.ClientSession() as session:
        tasks = [process_module(session, module) for module in modules]
        results = await asyncio.gather(*tasks)

    for module, filename, new_last_id in results:
        if filename:
            filenames.append(filename)
        if new_last_id is not None:
            last_ids[module] = new_last_id

    try:
        combined_workbook = await create_combined_workbook(filenames)
        if combined_workbook:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            combined_filename = f"Repsly_Export_Combined_{timestamp}.xlsx"
            combined_workbook.save(combined_filename)
            logger.info(f"Combined workbook saved as {combined_filename}")
        else:
            logger.error("Failed to create combined workbook.")
    except Exception as e:
        logger.error(f"Error creating or saving combined workbook: {str(e)}", exc_info=True)

    save_last_ids(last_ids)
    logger.info("Repsly data export completed.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Run Repsly data export modules.")
    parser.add_argument('modules', nargs='*', help='Modules to run. If none specified, all modules will run.')
    args = parser.parse_args()

    asyncio.run(main(args.modules))

